# -*- coding: utf-8 -*-
#!/usr/bin/python
import os
import datetime
from openpyxl import load_workbook, Workbook

BASE_PATH = os.path.dirname(os.path.abspath(__file__))


class ExcelUtil:
    '''
    excel操作类
    '''

    def __init__(self, path=None, index_table=0):
        '''
        :param path(str):文件路径，
        1. 为None时：调用save()方法时需要传入filename
        2. 如果不为空表示打开已有文件
        :param index_table(int):excel中的那个单元
        '''
        if path:
            # self.wb = load_workbook(os.path.join(BASE_PATH,path))
            self.wb = load_workbook(path)
        else:
            self.wb = Workbook()
        self.path = path
        sheets = self.wb.sheetnames
        sheet = sheets[index_table]
        self.sheet = self.wb[sheet]
        self.cell = self.sheet.cell

    def set_value_by_cell(self, row, column, value):
        '''
        通过cell设置值
        :param row(int):行
        :param column(int):列
        :param value(str):设置值
        '''
        self.cell(row,column).value = value

    def set_value_by_table(self, tag, value):
        '''
        通过A1坐标设置值
        例如：设置A1的值为hello
        set_value_by_table('A1', 'hello')
        :param tag(str):具体坐标
        :param value(str):值
        '''
        self.sheet[tag] = value

    def get_value_by_table(self, tag):
        '''
        通过A1坐标获取值
        例如：获取A1的值为hello
        get_value_by_table('A1')
        :param tag(str):具体坐标
        :param value(str):值
        :return <Cell 'Sheet'.A1>对象
        '''
        return self.sheet[tag].value

    def get_value_by_cell(self, row, column):
        '''
        通过cell坐标设置值 row行 colunm列
        例如：获取第一行第一列的值
        get_value_by_cell(self,1,1)
        :param row(str):行
        :param column(str):列
        '''
        return self.cell(row, column).value

    def get_max_row(self):
        '''获取最大行数 '''
        return self.sheet.max_row

    def get_max_col(self):
        '''获取最大列数 '''
        return self.sheet.max_column

    def get_col_value(self, column, row_start=1, row_end=None):
        """
        获取某列多少行的值，默认为所有
        introduce
        :param column(int):第几列
        :param row_start(int):开始的行，默认第一行
        :param row_end(int):结束的行，默认获取全部
        :return list()
        """
        if not row_end:
            row_end = self.get_max_row()
        column_data = []
        for i in range(row_start, row_end + 1):
            cell_value = self.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row, col_start=1, col_end=None):
        """
        获取某行多少列的值，默认为所有
        introduce
        :param row(int):第几行
        :param col_start(int):开始的列，默认第一列
        :param col_end(int):结束的列，默认获取全部
        :return list()
        """
        if not col_end:
            col_end = self.get_max_col()
        row_data = []
        for i in range(col_start, col_end + 1):
            cell_value = self.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    def save(self,filename=None):
        """
        获取文件名
        :param filename(str):保存的文件名
        :return:bool
        """
        if filename is None and self.path is None:
            print("保存失败：没有设置文件名")
            return False
        self.wb.save((filename if filename.endswith(".xlsx") else filename + ".xlsx") if filename else self.path)
        return True


# if __name__ == '__main__':
#     # 示例1：
#     # 1. 创建新文件在左边A1中写入数据
#     # 2. 文件名为test
#     ex = ExcelUtil()
#     ex.set_value_by_table("A1", '微信关注-给点知识公众号')
#     ex.set_value_by_table("A2", 1)
#     ex.save('test')
#     # 示例2
#     # 1. 打开文件test：
#     # 1. 获取A1中的数据
#     ex = ExcelUtil(path=r"test.xlsx")
#     a1 = ex.get_value_by_table("A1")
#     print(a1)

